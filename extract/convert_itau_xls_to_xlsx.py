#!/usr/bin/env python3
"""Convert Itau .xls bank files to .xlsx for tap-spreadsheets."""

from __future__ import annotations

import glob
import os
import sys
from datetime import date, time

import openpyxl
import xlrd


SOURCE_GLOB = "data/itau/debito/debito_xls/*.xls"
TARGET_DIR = "data/itau/debito/debito_xlsx"


def _convert_cell(book: xlrd.book.Book, cell: xlrd.sheet.Cell):
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return None
    if cell.ctype == xlrd.XL_CELL_TEXT:
        return cell.value
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        if float(cell.value).is_integer():
            return int(cell.value)
        return float(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        dt = xlrd.xldate_as_datetime(cell.value, book.datemode)
        if dt.time() == time(0, 0):
            return date(dt.year, dt.month, dt.day)
        return dt
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return None
    return cell.value


def convert_xls_to_xlsx(source_path: str, target_path: str) -> None:
    book = xlrd.open_workbook(source_path, formatting_info=False)
    out_wb = openpyxl.Workbook()

    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    for sheet_name in book.sheet_names():
        in_sheet = book.sheet_by_name(sheet_name)
        out_sheet = out_wb.create_sheet(title=sheet_name[:31])

        for row_idx in range(in_sheet.nrows):
            row_values = []
            for col_idx in range(in_sheet.ncols):
                cell = in_sheet.cell(row_idx, col_idx)
                row_values.append(_convert_cell(book, cell))
            out_sheet.append(row_values)

    os.makedirs(os.path.dirname(target_path), exist_ok=True)
    out_wb.save(target_path)


def main() -> int:
    source_files = sorted(glob.glob(SOURCE_GLOB))
    if not source_files:
        print(f"No source files found with pattern: {SOURCE_GLOB}")
        return 1

    os.makedirs(TARGET_DIR, exist_ok=True)
    converted = 0
    skipped = 0

    for source in source_files:
        base_name = os.path.splitext(os.path.basename(source))[0]
        target = os.path.join(TARGET_DIR, f"{base_name}.xlsx")

        if os.path.exists(target):
            src_mtime = os.path.getmtime(source)
            dst_mtime = os.path.getmtime(target)
            if dst_mtime >= src_mtime:
                skipped += 1
                continue

        convert_xls_to_xlsx(source, target)
        converted += 1

    print(f"Converted: {converted} | Skipped (up-to-date): {skipped}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
